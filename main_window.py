# _*_ coding: utf-8 _*_

import tkinter as tk
import tkinter.filedialog as fd
import tkinter.messagebox as mb
from tkcalendar import DateEntry

import openpyxl

from database import extract_dict

'''
Интерфейс и работа с ним
'''

NUM_COLUMN: str = 'A'
VOL_COLUMN: str = 'K'
DESC_COLUMN: str = 'L'

NUM_CELL: str = 'A1'
FIAS_CELL: str = 'I1'
DESC_CELL: str = f'{DESC_COLUMN}1'

NUM_INDEX: int = 1
FIAS_INDEX: int = 9
VOLUME_INDEX: int = 11
DESC_INDEX: int = 12


class MainWindow:
    '''
    Класс главного окна
    '''

    def __init__(self) -> None:
        '''
        Конструктор
        '''

        self.root = tk.Tk()
        self.root.title('Выбор даты и файла')

        # Создаем фрейм для первой строки
        top_frame = tk.Frame(self.root)
        top_frame.pack()

        # Создаем два элемента ввода даты
        date_label_start = tk.Label(top_frame, text='Дата начала:')
        date_label_start.pack(side=tk.LEFT)
        self.date_entry_start = DateEntry(top_frame, width=12, background='darkblue',
                                          foreground='white', borderwidth=2, locale='ru_RU')
        self.date_entry_start.pack(side=tk.LEFT, padx=5)
        date_label_end = tk.Label(top_frame, text='Дата конца:')
        date_label_end.pack(side=tk.LEFT)
        self.date_entry_end = DateEntry(top_frame, width=12, background='darkblue',
                                        foreground='white', borderwidth=2, locale='ru_RU')
        self.date_entry_end.pack(side=tk.LEFT, padx=5)

        # Создаем кнопку 'Выбрать файл' во второй строке
        bottom_frame = tk.Frame(self.root)
        bottom_frame.pack(pady=10)
        select_button = tk.Button(
            bottom_frame, text='Выбрать файл', command=self.__select_file)
        select_button.pack()

    def run(self) -> None:
        '''
        Запуск основного цикла
        '''

        self.root.mainloop()

    def __is_sheet_valid(self, sheet) -> bool:
        '''
        Проверка валидности листа
        sheet - лист excel
        '''

        if not str(sheet[NUM_CELL].value).startswith('№'):
            return False

        if not str(sheet[FIAS_CELL].value) == 'ФИАС':
            return False

        if not str(sheet[DESC_CELL].value) == 'Расшифровка':
            return False

        return True

    def __clean_sheet(self, worksheet) -> None:
        '''
        Очистить лист
        '''

        for row in worksheet.iter_rows(min_row=2, min_col=VOLUME_INDEX, max_col=DESC_INDEX):
            for cell in row:
                # Присваевается нулевое значение ячейке
                cell.value = None

    def __handle_sheet(self, worksheet, dictionary) -> None:
        '''
        Обработать лист
        worksheet - лист
        dictionary - словарь
        '''

        # Подготовка листа к заполнению
        self.__clean_sheet(worksheet)

        # Итоговая сумма
        total_sum: float = 0

        # Суммы расширений
        total_values: dict = dict()

        # Заполнение листа
        for row in worksheet.iter_rows(min_row=2, min_col=FIAS_INDEX, max_col=DESC_INDEX):
            # Пропускаем пустые ФИАСы
            if not row[0].value or not dictionary.__contains__(row[0].value):
                continue

            # Итоговая сумма для ФИАСа
            current_total: float = 0

            # Обходим по итератору словарь с ФИАСами
            for key, val in dictionary[row[0].value]:
                if not total_values.__contains__(key):
                    total_values[key]: float = 0

                total_values[key] += val
                current_total += val

            # Заполняем ячейку объёма
            row[2].value = current_total

            # Заполняем ячейку расшифровки
            row[3].value = str(dictionary[row[0].value])

            # Считаем итоговую сумму
            total_sum += current_total

        # Поиск последней строки в первом столбце, идём с конца "А" столбца по всем ячейкам
        # Этот костыль необходимое зло
        num_rows = worksheet.max_row
        for row in range(num_rows, NUM_INDEX, -1):
            if worksheet.cell(row=row, column=NUM_INDEX).value:
                num_rows = row + 1
                break

        # Запись итогового результата
        worksheet[f'{VOL_COLUMN}{num_rows}'].value = total_sum
        worksheet[f'{DESC_COLUMN}{num_rows}'].value = 'Итого'

        # Запись суммарных значений атрибутов
        num_rows += 1
        count: int = 0

        for key, val in total_values.items():
            worksheet[f'{VOL_COLUMN}{num_rows + count}'].value = val
            worksheet[f'{DESC_COLUMN}{num_rows + count}'].value = key
            count += 1

    def __handle_dict(self, dictionary, workbook) -> None:
        '''
        Обработать словарь
        dictionary - словарь
        path - путь до файла
        '''

        # Обходим все листы в файле
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            # Если лист не подходит критериям, пропускаем его
            if not self.__is_sheet_valid(worksheet):
                continue

            # Обработка листа
            self.__handle_sheet(worksheet, dictionary)

    def __select_file(self) -> None:
        '''
        Обработка нажатия кнопки
        '''

        # Если дата конца интервала позже даты начала, выдаём ошибку
        if self.date_entry_start.get_date() > self.date_entry_end.get_date():
            start: str = str(self.date_entry_start.get_date()
                             ).replace('-', '.')
            end: str = str(self.date_entry_end.get_date()).replace('-', '.')
            error_text: str = f'Дата начала {start} больше даты конца {end}'
            mb.showerror('Ошибка', error_text)
            return

        try:
            path: str = fd.askopenfilename()

            # Если был отменён ввод, пропускаем
            if not path:
                return

            # Открываем для записи с поддержкой макросов для корректного сохранения файлов
            workbook = openpyxl.load_workbook(
                path, read_only=False, keep_vba=True)

            start: str = str(self.date_entry_start.get_date()
                             ).replace('-', '.')
            end: str = str(self.date_entry_end.get_date()).replace('-', '.')
            dictionary = extract_dict(start, end)

            self.__handle_dict(dictionary, workbook)

            workbook.save(path)
            mb.showinfo('Уведомление', 'Обработка файла завершена успешно!')

        except Exception as e:
            # Вывод текста исключения как ошибки
            mb.showerror('Ошибка', e)
            print(e)
