import pyodbc
import openpyxl
import tkinter as tk
from tkinter import filedialog
from tkcalendar import DateEntry

CONNECTION: str = 'Driver={SQL Server};Server=10.1.1.7;Uid=dudnik; Pwd=tvk2022dud;'


XLSM_FILE: str = 'C:/Users/nepri/Рабочий стол/Водный баланс в. 2.xlsm'


def is_sheet_valid(sheet) -> bool:
    return sheet['A1'].value == '№\nп/п'


def connect() -> None:
    # con = pyodbc.connect(CONNECTION)

    # cursor = con.cursor()
    # for row in cursor.tables():
    #     print(row.table_name)

    # data_begin: str = '01.09.2022'
    # data_end: str = '02.09.2022'
    # call: str = f'EXEC Get_Obem_Doma @Data_Begin="{data_begin}", @Data_End="{data_end}"'

    # rows = cursor.execute(call)
    # with open(file='file.txt', mode='w') as f:
    #     for row in rows:
    #         f.write(str(row))
    #         f.write('\n')
    #         print(row)

    workbook = openpyxl.load_workbook(XLSM_FILE)
    worksheet = workbook.active

    # for row in worksheet.iter_rows(min_row=1, max_row=1, max_col=10):
    #     for cell in row:
    #         print(cell.value)

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]

        if not is_sheet_valid(worksheet):
            continue

        print(f'Sheet \'{sheet_name}\' is valid')


def select_file():
    file_path = filedialog.askopenfilename()
    print("Выбранный файл: ", file_path)


def create_window() -> None:
    root = tk.Tk()
    root.title("Выбор даты и файла")

    # Создаем фрейм для первой строки
    top_frame = tk.Frame(root)
    top_frame.pack()

    # Создаем два элемента ввода даты
    date1_label = tk.Label(top_frame, text="Дата 1:")
    date1_label.pack(side=tk.LEFT)
    date1_entry = DateEntry(top_frame, width=12, background='darkblue',
                            foreground='white', borderwidth=2, locale='ru_RU')
    date1_entry.pack(side=tk.LEFT, padx=5)
    date2_label = tk.Label(top_frame, text="Дата 2:")
    date2_label.pack(side=tk.LEFT)
    date2_entry = DateEntry(top_frame, width=12, background='darkblue',
                            foreground='white', borderwidth=2, locale='ru_RU')
    date2_entry.pack(side=tk.LEFT, padx=5)

    # Создаем кнопку "Выбрать файл" во второй строке
    bottom_frame = tk.Frame(root)
    bottom_frame.pack(pady=10)
    select_button = tk.Button(
        bottom_frame, text="Выбрать файл", command=select_file)
    select_button.pack()

    root.mainloop()


if __name__ == '__main__':
    try:
        create_window()
        # connect()
    except Exception as e:
        print(e)
